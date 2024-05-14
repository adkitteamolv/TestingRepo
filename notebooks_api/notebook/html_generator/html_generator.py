"""
    This class is used to generate html from files like ipynb ,py and R
"""
import os
import base64
import logging
import csv
from nbconvert import HTMLExporter
import nbformat
import markdown
from tabulate import tabulate
import openpyxl
from pydocx import PyDocX
from .lang_to_html import hilite_me

# pylint: disable=invalid-name
log = logging.getLogger("notebooks_api")

# pylint: disable=useless-object-inheritance

class HtmlGenerator(object):
    """
    class decleration for html generator
    """
    def __init__(self, content, ext):
        self.ext = ext
        if self.ext in ['.png', '.jpeg', '.jpg', '.bmp', '.gif']:
            self.content = content
        else:
            self.content = base64.decodebytes(content.encode())
        self.fields = {'.ipynb': self.ipynb_to_html, '.md': self.convert_md_to_html,
                       '.csv' : self.convert_csv_to_html, '.xlsx': self.convert_xlsx_to_html,
                       '.docx' : self.convert_docx_to_html, '.png': self.embed_image_to_html,
                       '.jpg': self.embed_image_to_html, '.jpeg': self.embed_image_to_html,
                       '.bmp': self.embed_image_to_html, '.gif': self.embed_image_to_html}

    def convert_file_to_html(self):
        """
        this method is orchestration to call method based on the extension match
        :return:
        """
        log.debug("---- File extension: %s ----", self.ext)
        if not self.ext:
	        self.ext = '.txt'
        return self.fields.get(self.ext, self.other_language_to_html)()

    def other_language_to_html(self):
        """
        this method renders language scripts to html
        :return:
        """
        try:
            html = hilite_me(self.content, self.ext)
            return html
        except Exception as ex:
            log.debug("---- Script Preview error: %s ----", ex)
            raise Exception(ex)

    def convert_md_to_html(self):
        """
        this method is used convert .md to html
        :return:
        """
        content_text = '\n'.join(map(str.rstrip, self.content.decode().split("\\n")))
        body = markdown.markdown(content_text)
        html = ('<html><head>'+
                '<style> *{font-family: Inter, Lato, sans-serif !important}' +
                'code { white-space: pre-wrap; word-wrap: normal; ' +
                'word-break: keep-all; padding: 2px 6px;' +
                'color: #3095d9; background-color: rgba(48,149,217, 0.1);' +
                'border-radius: 4px; }' +
                'a { color: #2468a4; text-decoration: none; }'+
                'a:hover { text-decoration: underline; }</style></head><body>' +
                body +
                '</body></html>')
        return html

    def ipynb_to_html(self):
        """
        this method is used convert ipynb to html
        :return:
        """
        html_exporter = HTMLExporter(template_name='lab')
        file_content = nbformat.reads(self.content, as_version=4)
        # pylint: disable=unused-variable
        (body, resources) = html_exporter.from_notebook_node(file_content)
        return body

    def convert_csv_to_html(self):
        """
        this method is used to convert csv to html
        :return:
        """
        record = self.content.decode().splitlines()
        read_csv = csv.reader(record)
        list_csv = []
        for line in read_csv:
            list_csv.append(line)
        table = tabulate(list_csv[1:], headers=list_csv[0], tablefmt='html')
        dir_path = os.path.dirname(os.path.realpath(__file__))
        style = open(dir_path + "/style_for_table.txt", "r").read()
        html = style + table + '</body></html>'
        return html

    def convert_xlsx_to_html(self):
        """
        this method is used to convert xlsx to html
        :return:
        """
        # writing the content to a binary file
        dir_path = "/tmp"
        with open(dir_path + "/sample.xlsx", "wb") as binary_file:
            # Write bytes to file
            binary_file.write(self.content)
        binary_file.close()
        # reading the xlsx
        wb = openpyxl.load_workbook(dir_path + "/sample.xlsx")
        ws = wb[wb.sheetnames[0]]
        list_excel = []
        for cells in ws.iter_rows():
            list_excel.append([cell.value for cell in cells])
        table = tabulate(list_excel[1:], headers=list_excel[0], tablefmt='html')
        dir_path = os.path.dirname(os.path.realpath(__file__))
        style = open(dir_path + "/style_for_table.txt", "r").read()
        html = style + table + '</body></html>'
        return html

    def convert_docx_to_html(self):
        """
        This method converts docx to html
        :return:
        """
        # writing the content to a binary file
        dir_path_temp = "/tmp"
        with open(dir_path_temp + "/sample.docx", "wb") as binary_file:
            # Write bytes to file
            binary_file.write(self.content)
        binary_file.close()
        html = PyDocX.to_html(open(dir_path_temp + "/sample.docx", 'rb'))
        return html

    def embed_image_to_html(self):
        """
        This embeds image to html
        :return:
        """
        html = '<img src="data:image/'+self.ext+';base64,'\
               +self.content+'"width="100%" height="100%" />'
        return html
